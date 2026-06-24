from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from models import Task, TaskStatus, TaskLog


def generate_weekly_timesheet(session: Session, year: int, week: int, output_path: str) -> None:
    wb = Workbook()

    ws = wb.active
    assert isinstance(ws, Worksheet), "Не удалось инициализировать рабочий лист Excel"

    ws.title = "Все задачи" if (year == 0 and week == 0) else f"Неделя {week}"

    headers = ["ID", "Заголовок", "Исполнитель", "Статус", "Приоритет", "Дедлайн", "Дата закрытия"]
    ws.append(headers)

    tasks = session.query(Task).all()
    # Сортируем: сначала по критичности приоритета (weight), затем по дедлайну (ближайшие выше)
    tasks = sorted(tasks, key=lambda t: (t.priority, -t.deadline.timestamp()), reverse=True)

    row_idx = 2
    for task in tasks:
        if year != 0 and week != 0:
            task_year, task_week, _ = task.deadline.isocalendar()
            if task_year != year or task_week != week:
                continue

        closed_log = session.query(TaskLog).filter(
            TaskLog.task_id == task.id,
            TaskLog.new_status == TaskStatus.DONE
        ).order_by(TaskLog.changed_at.desc()).first()

        deadline_val = task.deadline
        closed_at_val = closed_log.changed_at if closed_log else ""

        ws.append([
            task.id,
            task.title,
            task.assignee.fullname if task.assignee else "Не назначен",
            task.status.value,
            task.priority.value,
            deadline_val,
            closed_at_val
        ])

        ws[f"F{row_idx}"].number_format = 'yyyy-mm-dd hh:mm'
        if closed_at_val:
            ws[f"G{row_idx}"].number_format = 'yyyy-mm-dd hh:mm'

        row_idx += 1

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    done_val = TaskStatus.DONE.value

    for r in range(2, row_idx):
        # Выносим создание правил на одну строку и глушим предупреждения openpyxl прямо здесь
        rule_red = FormulaRule(formula=[f'AND($D${r}<>"{done_val}", $F${r}<NOW())'],
                               fill=red_fill)  # type: ignore[no-untyped-call]
        rule_green = FormulaRule(formula=[f'AND($D${r}="{done_val}", $G${r}<$F${r})'],
                                 fill=green_fill)  # type: ignore[no-untyped-call]
        rule_yellow = FormulaRule(formula=[f'AND($D${r}<>"{done_val}", $F${r}>NOW(), $F${r}<=NOW()+1)'],
                                  fill=yellow_fill)  # type: ignore[no-untyped-call]

        # Добавляем уже созданные объекты правил в условное форматирование
        ws.conditional_formatting.add(f"A{r}:G{r}", rule_red)
        ws.conditional_formatting.add(f"A{r}:G{r}", rule_green)
        ws.conditional_formatting.add(f"A{r}:G{r}", rule_yellow)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)

        col_index = col[0].column
        assert isinstance(col_index, int), "Индекс колонки должен быть целым числом"

        col_letter = get_column_letter(col_index)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(output_path)
